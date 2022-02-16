from bs4 import BeautifulSoup
import requests
import lxml
import openpyxl

todoslosImpuestos = 1.66 #taxes summed up

#Returns: [name, price, price with taxes, str of prices and names]
def scrapingSteam(argument):
    if (argument.startswith("https://store.steampowered.com/app") == -1):
       print("Error, ese link no se puede procesar, solo los de https://store.steampowered.com/app son utilizables")
    else:
        r = requests.get(argument)
        soup = BeautifulSoup(r.content, "lxml")
        name = soup.find("div", class_="apphub_AppName").text
        prices = ""

        #First of all: we get all the prices
        for x in soup.find_all("div", class_="game_purchase_action_bg"):
            prices = f"{prices} ~ {x.get_text()}"

        #We do this in order to be sure it is the price we want because prices contains all the prices in the page
        prices = prices.strip()
        prices = prices.split("~")
        firstPrice = prices[1].strip()

        #This line adds compatibility with packages
        if (firstPrice.find("Package info") > -1):
            firstPrice = prices[2].strip()

        #If it can find free: its free
        if (firstPrice.find("Free") > -1):
            price = 0
            priceWithTaxes = 0

        #If it can find download: then it is a demo
        elif (firstPrice.find("Download") > -1):
            price = 0
            priceWithTaxes = 0

        #if you can put it in your cart, then it is a paid game 
        elif (firstPrice.find("Add to Cart") > -1):

            #Getting the last price in the first element of the list, works with games on a discount
            firstPrice = firstPrice.split("$")
            firstPrice = firstPrice[-1]

            #Now, to avoid a million methods() we'll check if its a character we want and add it to a string
            price = ""
            for x in firstPrice:
                if x.isdigit() == True: price += x
                elif x == ",": price += "."
                    
            #Finally, we get a float and then the taxes are calculated
            priceWithTaxes = float(price) * todoslosImpuestos

        else:
            price = 0
            priceWithTaxes = 0
            #If it we can't buy it, to avoid errors, its worth 0.

        return([ name, price, round(priceWithTaxes, 2), (f"{name}: {round(priceWithTaxes, 2)} \n"), round(priceWithTaxes, 2)])

#Default cart, output is on console
def cart():

    listFromInput = scrapingSteam(input("Ingresá el link!: "))
    nameWithPrice = listFromInput[3]
    priceTaxed = listFromInput[2]

    #Now, the input asks us to either enter a new link or to end
    ifStatement = input("Deseas añadir algo más al carrito? Y o N: ")
    while (ifStatement == "Y" or ifStatement == "y"):
        listFromInput = (scrapingSteam((input("Ingresá el link!: "))))
        nameWithPrice += listFromInput[3]
        priceTaxed += listFromInput[2]
        ifStatement = input("Deseas añadir algo más al carrito? Y o N: ")
    print(nameWithPrice + "Total:", round(priceTaxed, 2))

#Writes on exel
def cartExel():
    listFromInput = scrapingSteam(input("Ingresá el link!: "))
    nameWithPrice = listFromInput[3]
    priceTaxed = listFromInput[2]
    priceWithoutTaxes = listFromInput[1]
    #It starts same as before but we have an extra variable, the original untaxed price

    my_wb = openpyxl.Workbook()
    my_sheet = my_wb.active
    my_row = 3
    #Now, a spreadsheet is created and named, also my_row is initialized 

    my_sheet.cell(row=1, column=1).value = "Juego"
    my_sheet.cell(row=1, column=2).value = "Costo sin impuestos"
    my_sheet.cell(row=1, column=3).value = "Costo con impuestos"

    for x in range(1, 4):
        my_sheet.cell(row=my_row, column=x).value = listFromInput[x-1]

    ifStatement = input("Deseas añadir algo más al carrito? Y o N: ")

    while (ifStatement == "Y" or ifStatement == "y"):
        listFromInput = (scrapingSteam((input("Ingresá el link!: "))))
        nameWithPrice += listFromInput[3]
        priceTaxed += listFromInput[2]
        priceWithoutTaxes += listFromInput[1]
        #this is the same as before, only that one more variable is stored

        for x in range(1, 4):
            my_sheet.cell(row=my_row, column=x).value = listFromInput[x - 1] #nombre
        my_row += 2
        #Now we write the values from the list into the xlsx file, also adding +1 to the row so it does not overwrite on the next link

        ifStatement = input("Deseas añadir algo más al carrito? Y o N: ")

    my_sheet.cell(row=my_row, column=1).value = "Total" #nombre
    my_sheet.cell(row=my_row, column=2).value = priceWithoutTaxes #costo SIN impuestos
    my_sheet.cell(row=my_row, column=3).value = priceTaxed #costo CON impuestos

    my_wb.save("CarritoDeSteam.xlsx")

    print(nameWithPrice + "Total: " , round(priceTaxed, 2))

def start():
    ifStatement = input("Deseas escribirlo en un documento de exel con más detalles? Y o N: ")
    if (ifStatement == "N" or ifStatement == "n"):   cart()
    elif (ifStatement == "Y" or ifStatement == "y"):   cartExel()
#Running input to make a choice.

start()
